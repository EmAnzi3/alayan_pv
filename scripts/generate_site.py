from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime, date
import json, re, shutil

HEADER_ROW = 4

CSS = """
:root{
  --bg:#f4efe7; --card:#fbf8f3; --line:#e2d8c7; --text:#2f2a24; --muted:#6d655d;
  --gold:#c98318; --blue:#245e8a; --green:#2b7a69; --chip:#efe4cf; --olive:#6a8f2a;
  --violet:#a06cd5; --brown:#8f5f3f; --neutral:#8b8378;
}
*{box-sizing:border-box}
body{margin:0;font-family:Inter,Segoe UI,Arial,sans-serif;background:linear-gradient(180deg,#f5f1ea 0%,#f0e9dd 100%);color:var(--text)}
.wrap{max-width:1280px;margin:0 auto;padding:24px 24px 40px}
.top-link{display:inline-flex;align-items:center;gap:8px;margin:0 0 14px;color:var(--blue);text-decoration:none;font-weight:700}
.header{display:flex;justify-content:space-between;align-items:flex-start;gap:16px;margin-bottom:18px}
.h1{font-size:52px;line-height:1.02;font-weight:800;letter-spacing:-.03em;margin:0}
.sub{margin-top:8px;color:var(--muted);font-size:17px}
.badge{padding:8px 12px;border-radius:12px;background:#f4e5c8;color:var(--gold);font-weight:700;font-size:13px}
.cards{display:grid;grid-template-columns:repeat(4,1fr);gap:16px;margin:22px 0}
.card{background:rgba(255,255,255,.68);backdrop-filter:blur(4px);border:1px solid var(--line);border-radius:22px;padding:18px 18px 16px;box-shadow:0 10px 30px rgba(80,60,20,.05)}
.kpi-label{font-size:12px;font-weight:700;color:var(--muted);text-transform:uppercase;letter-spacing:.08em}
.kpi-value{font-size:44px;font-weight:800;line-height:1;margin:8px 0 6px}
.kpi-sub{font-size:14px;color:var(--muted)}
.kpi-value.gold{color:var(--gold)} .kpi-value.green{color:var(--green)}
.panel-grid{display:grid;grid-template-columns:1fr 1.35fr;gap:16px;margin-bottom:16px}
.panel{background:rgba(255,255,255,.68);border:1px solid var(--line);border-radius:22px;padding:16px 16px 14px;box-shadow:0 10px 30px rgba(80,60,20,.05)}
.panel h3{margin:0 0 12px;font-size:18px;letter-spacing:.02em}
.flex{display:flex;gap:18px;align-items:center}
.donut-wrap{width:220px;height:220px;display:grid;place-items:center;position:relative}
.donut{width:190px;height:190px;border-radius:50%;background:conic-gradient(var(--gold) 0 100%);position:relative}
.donut::after{content:'';position:absolute;inset:32px;background:#fffaf3;border-radius:50%;box-shadow:inset 0 0 0 1px var(--line)}
.donut-center{position:absolute;text-align:center;z-index:2}
.donut-center .big{font-size:30px;font-weight:800}
.legend{display:grid;gap:10px;min-width:220px}
.legend-row{display:grid;grid-template-columns:16px 1fr auto;gap:10px;align-items:center;font-size:15px}
.dot{width:12px;height:12px;border-radius:4px}
.filters{display:flex;gap:12px;flex-wrap:wrap;margin:8px 0 16px}
.select,.jump-select{background:#fff;border:1px solid var(--line);border-radius:14px;padding:12px 14px;min-width:210px;color:var(--text);font-size:15px}
.jump{display:flex;gap:10px;flex-wrap:wrap;align-items:center}
.jump button{border:0;background:var(--gold);color:#fff;padding:12px 16px;border-radius:14px;font-weight:700;cursor:pointer}
.note{background:rgba(255,255,255,.58);border:1px dashed var(--line);border-radius:18px;padding:14px 16px;margin:0 0 16px;color:var(--text)}
.bars{display:grid;gap:10px;margin-top:8px}
.bars.scrollable{max-height:320px;overflow-y:auto;padding-right:8px}
.bars.scrollable::-webkit-scrollbar,.table-scroll::-webkit-scrollbar,.gantt-scroll::-webkit-scrollbar{width:10px;height:10px}
.bars.scrollable::-webkit-scrollbar-thumb,.table-scroll::-webkit-scrollbar-thumb,.gantt-scroll::-webkit-scrollbar-thumb{background:#dbcaa8;border-radius:999px}
.bars.scrollable::-webkit-scrollbar-track,.table-scroll::-webkit-scrollbar-track,.gantt-scroll::-webkit-scrollbar-track{background:#f1e8d8;border-radius:999px}
.bar-row{display:grid;grid-template-columns:130px 1fr 80px;gap:12px;align-items:center}
.bar-row.linkable{cursor:pointer}
.bar-label-link{color:var(--blue);text-decoration:none;font-weight:700}
.bar-track{height:18px;background:#f1e8d8;border-radius:999px;overflow:hidden;position:relative}
.bar-fill{height:100%;background:linear-gradient(90deg,var(--gold),#e2a23d);border-radius:999px}
.table-card{background:rgba(255,255,255,.68);border:1px solid var(--line);border-radius:22px;padding:16px 16px 8px;box-shadow:0 10px 30px rgba(80,60,20,.05);margin-top:16px}
.table-head{display:flex;justify-content:space-between;align-items:center;gap:16px;margin-bottom:12px}
.table-title{font-size:18px;font-weight:800}
.small{font-size:13px;color:var(--muted)}
.table-scroll{max-height:520px;overflow:auto;padding-right:6px;border-radius:16px}
table{width:100%;min-width:980px;border-collapse:separate;border-spacing:0 8px}
thead th{text-align:left;font-size:12px;letter-spacing:.08em;text-transform:uppercase;color:var(--muted);padding:0 8px 4px}
tbody td{background:#fffdf9;padding:12px 8px;border-top:1px solid #f0e7d8;border-bottom:1px solid #f0e7d8;font-size:14px}
tbody tr td:first-child{border-left:1px solid #f0e7d8;border-radius:12px 0 0 12px}
tbody tr td:last-child{border-right:1px solid #f0e7d8;border-radius:0 12px 12px 0}
.chip{display:inline-block;padding:6px 10px;border-radius:999px;font-weight:700;font-size:12px}
.aut{background:#f4e5c8;color:var(--gold)} .proc{background:#dceaf5;color:var(--blue)} .costr{background:#dceee7;color:var(--green)} .comm{background:#e9efd7;color:var(--olive)}
.std{background:#efe7fb;color:var(--violet)} .prog{background:#efe1d8;color:var(--brown)}
.state-pipeline{background:#ece8e2;color:#534b43} .state-attivo{background:#dceee7;color:var(--green)} .state-completato{background:#e9efd7;color:var(--olive)}
.progress{display:flex;align-items:center;gap:10px}
.progress>span{min-width:42px;font-weight:700}
.progress .track{height:10px;width:110px;background:#efe4d2;border-radius:999px;overflow:hidden}
.progress .fill{height:100%;background:linear-gradient(90deg,#e2a13f,var(--gold));border-radius:999px}
.gantt-shell{background:rgba(255,255,255,.68);border:1px solid var(--line);border-radius:22px;padding:16px;box-shadow:0 10px 30px rgba(80,60,20,.05);margin-top:16px}
.gantt-scroll{overflow:auto;border-radius:16px}
.gantt{min-width:1000px}
.gantt-header,.gantt-row{display:grid;grid-template-columns:280px 1fr;gap:12px;align-items:center}
.gantt-header{padding:0 0 8px}
.gantt-label{font-size:12px;letter-spacing:.08em;text-transform:uppercase;color:var(--muted);font-weight:700}
.gantt-timeline{position:relative;display:grid;grid-auto-flow:column;grid-auto-columns:minmax(70px,1fr);gap:0;border-left:1px solid #eadfce}
.gantt-month{font-size:12px;color:var(--muted);padding:0 8px 6px;border-right:1px solid #eadfce}
.gantt-row{padding:10px 0;border-top:1px solid #efe5d5}
.gantt-name{font-size:14px;font-weight:700}
.gantt-sub{font-size:12px;color:var(--muted);margin-top:4px}
.gantt-track{position:relative;height:32px;background:repeating-linear-gradient(90deg,#faf4ea 0,#faf4ea calc(100%/var(--months)),#f4ead8 calc(100%/var(--months)),#f4ead8 calc(100%/var(--months) + 1px));border-radius:12px;overflow:hidden}
.gantt-seg{position:absolute;height:16px;top:8px;border-radius:999px;box-shadow:0 1px 0 rgba(255,255,255,.65) inset}
.gantt-empty{color:var(--muted);padding:12px 0}
.footer{margin-top:12px;color:var(--muted);font-size:13px}
@media (max-width:980px){
  .wrap{padding:22px 18px 32px}
  .header{flex-direction:column;gap:14px;align-items:flex-start}
  .cards{grid-template-columns:1fr 1fr}
  .panel-grid{grid-template-columns:1fr}
  .table-head{flex-direction:column;align-items:flex-start}
  .filters,.jump{width:100%}
  .select,.jump-select{flex:1 1 220px;min-width:180px}
}
@media (max-width:700px){
  .wrap{padding:16px 12px 24px}
  .cards{grid-template-columns:1fr}
  .h1{font-size:34px}
  .sub{font-size:15px}
  .card,.panel,.table-card,.gantt-shell{padding:14px}
  .kpi-value{font-size:38px}
  .bar-row{grid-template-columns:92px 1fr 56px;gap:8px}
  .flex{flex-direction:column;align-items:flex-start}
  .donut-wrap{width:180px;height:180px}
  .donut{width:160px;height:160px}
  .donut::after{inset:28px}
  .legend{min-width:0;width:100%}
  .filters,.jump{display:grid;grid-template-columns:1fr;gap:10px;width:100%}
  .select,.jump-select{width:100%;min-width:0;padding:11px 12px;font-size:14px}
  .bars.scrollable{max-height:240px}
  .table-scroll{max-height:60vh}
  table{min-width:840px}
  .gantt{min-width:860px}
  .gantt-header,.gantt-row{grid-template-columns:220px 1fr}
}
"""

def slugify(s):
    s = (s or "").strip().lower()
    s = s.replace("à","a").replace("è","e").replace("é","e").replace("ì","i").replace("ò","o").replace("ù","u")
    s = re.sub(r"[^a-z0-9]+", "-", s)
    return s.strip("-") or "filiale"

def fmt_date_py(v):
    if v in (None, "", 0):
        return ""
    if isinstance(v, datetime):
        if v.year <= 1970:
            return ""
        return v.strftime("%d %b %Y")
    if isinstance(v, date):
        if v.year <= 1970:
            return ""
        return v.strftime("%d %b %Y")
    if isinstance(v, (int, float)):
        if v <= 0:
            return ""
        try:
            from openpyxl.utils.datetime import from_excel
            d = from_excel(v)
            if isinstance(d, datetime):
                d = d.date()
            if d.year <= 1970:
                return ""
            return d.strftime("%d %b %Y")
        except Exception:
            return ""
    s = str(v).strip()
    if not s or s.startswith("1970-01-01"):
        return ""
    return s

def normalize_text(v):
    if v is None:
        return ""
    return str(v).strip()

def normalize_num(v):
    if v in (None, ""):
        return 0.0
    try:
        return float(v)
    except Exception:
        return 0.0

def load_config():
    cfg_path = Path(__file__).with_name("config.json")
    with cfg_path.open("r", encoding="utf-8") as f:
        return json.load(f)

def read_sheet_rows(xlsx_path):
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["Progetti"]
    headers = [c.value for c in ws[HEADER_ROW]]
    records = []
    filiale_default = ""
    area_default = ""
    if "Istruzioni" in wb.sheetnames:
        iw = wb["Istruzioni"]
        filiale_default = normalize_text(iw["B4"].value)
        area_default = normalize_text(iw["B5"].value)
    for row in ws.iter_rows(min_row=HEADER_ROW+1, values_only=True):
        rec = {headers[i]: row[i] for i in range(len(headers))}
        filiale = normalize_text(rec.get("Filiale")) or filiale_default
        area = normalize_text(rec.get("Area")) or area_default
        progetto = normalize_text(rec.get("Progetto"))
        cliente = normalize_text(rec.get("Cliente"))
        provincia = normalize_text(rec.get("Provincia"))
        mwp = normalize_num(rec.get("Potenza MWp"))
        any_main = any([
            progetto, cliente, provincia,
            mwp not in (0, 0.0),
            normalize_text(rec.get("Stato progetto")),
            normalize_text(rec.get("Fase attuale")),
            normalize_text(rec.get("Note"))
        ])
        if not any_main:
            continue

        note = normalize_text(rec.get("Note"))
        ultimo = fmt_date_py(rec.get("Ultimo aggiornamento"))
        stato = normalize_text(rec.get("Stato progetto"))
        fase = normalize_text(rec.get("Fase attuale"))
        avanz = normalize_num(rec.get("Avanzamento %"))
        if avanz > 1:
            avanz = avanz / 100.0
        if avanz < 0:
            avanz = 0
        if avanz > 1:
            avanz = 1

        segments = []
        phase_pairs = [
            ("Studio di fattibilità", "SdF Inizio", "SdF Fine"),
            ("Progettazione", "Prog. Inizio", "Prog. Fine"),
            ("Autorizzazioni", "Aut. Inizio", "Aut. Fine"),
            ("Procurement", "Proc. Inizio", "Proc. Fine"),
            ("Costruzione", "Costr. Inizio", "Costr. Fine"),
            ("Commissioning", "Comm. Inizio", "Comm. Fine"),
        ]
        for label, c1, c2 in phase_pairs:
            d1 = rec.get(c1)
            d2 = rec.get(c2)
            if isinstance(d1, datetime): d1 = d1.date()
            if isinstance(d2, datetime): d2 = d2.date()
            if isinstance(d1, date) and isinstance(d2, date) and d1.year > 1970 and d2.year > 1970:
                if d2 < d1:
                    d1, d2 = d2, d1
                segments.append({
                    "label": label,
                    "start": d1.isoformat(),
                    "end": d2.isoformat(),
                })

        records.append({
            "filiale": filiale,
            "area": area,
            "cliente": cliente,
            "progetto": progetto,
            "provincia": provincia,
            "comune": normalize_text(rec.get("Comune")),
            "regione": normalize_text(rec.get("Regione")),
            "mwp": mwp,
            "stato": stato,
            "fase": fase,
            "ultimo_agg": ultimo,
            "note": note,
            "agrivoltaico": ("agrivoltaico" in note.lower()),
            "avanz": avanz,
            "segments": segments,
        })
    return records, filiale_default, area_default

def render_page(title, subtitle, mode, data_rows, branch_name="", branch_area="", all_branches=None):
    all_branches = all_branches or []
    page_data = json.dumps(data_rows, ensure_ascii=False)
    branches_data = json.dumps(all_branches, ensure_ascii=False)
    back_link = ''
    jump_block = ''
    if mode == "branch":
        back_link = '<a class="top-link" href="../index.html">← Torna alla overview nazionale</a>'
    else:
        jump_block = """
        <div class="table-card" style="margin-top:0;margin-bottom:16px">
          <div class="table-head">
            <div class="table-title">Vai alle filiali</div>
            <div class="jump">
              <select id="jumpBranch" class="jump-select"><option value="">Seleziona una filiale</option></select>
              <button id="jumpBtn" type="button">Apri dashboard</button>
            </div>
          </div>
        </div>
        """
    scope_title = "MWp per provincia" if mode == "branch" else "MWp per filiale"
    list_title = "Tutti i progetti" if mode == "branch" else "Progetti aggregati"
    scope_placeholder = "Tutte le province" if mode == "branch" else "Tutte le filiali"
    scope_id = "provFilter" if mode == "branch" else "filialeFilter"
    area_select = '' if mode == "branch" else '<select id="areaFilter" class="select"><option value="">Tutte le aree</option></select>'
    branch_back_footer = '<div class="footer">Utilizzare i filtri per visualizzare gli elementi d&apos;interesse.</div>'
    return f"""<!DOCTYPE html>
<html lang="it">
<head>
<meta charset="utf-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1"/>
<title>{title}</title>
<style>{CSS}</style>
</head>
<body>
<div class="wrap">
  {back_link}
  <div class="header">
    <div>
      <h1 class="h1">{title}</h1>
      <div class="sub">{subtitle}</div>
    </div>
    <div class="badge">Alayan</div>
  </div>

  {jump_block}

  <div class="cards">
    <div class="card"><div class="kpi-label">Progetti</div><div id="kpiProjects" class="kpi-value">0</div><div class="kpi-sub">totali in vista corrente</div></div>
    <div class="card"><div class="kpi-label">Potenza totale</div><div id="kpiMwp" class="kpi-value gold">0,0</div><div class="kpi-sub">MWp complessivi</div></div>
    <div class="card"><div class="kpi-label">In costruzione</div><div id="kpiBuild" class="kpi-value green">0</div><div class="kpi-sub">progetti attivi in cantiere</div></div>
    <div class="card"><div class="kpi-label">Agrivoltaico</div><div id="kpiAgri" class="kpi-value green">0</div><div class="kpi-sub">progetti agrivoltaici</div></div>
  </div>

  <div class="panel-grid">
    <div class="panel">
      <h3>Fase attuale</h3>
      <div class="flex">
        <div class="donut-wrap">
          <div class="donut-center"><div id="donutTotal" class="big">0</div><div class="small">progetti</div></div>
          <div id="donut" class="donut"></div>
        </div>
        <div id="phaseLegend" class="legend"></div>
      </div>
    </div>
    <div class="panel">
      <h3>{scope_title}</h3>
      <div id="scopeBars" class="bars {'scrollable' if mode=='global' else ''}"></div>
    </div>
  </div>

  <div class="note" id="insightBox"></div>

  <div class="table-card">
    <div class="table-head">
      <div class="table-title">{list_title}</div>
      <div class="filters">
        {area_select}
        <select id="{scope_id}" class="select"><option value="">{scope_placeholder}</option></select>
        <select id="faseFilter" class="select"><option value="">Tutte le fasi</option></select>
        <select id="stateFilter" class="select"><option value="">Tutti gli stati</option></select>
      </div>
    </div>
    <div class="table-scroll">
      <table>
        <thead>
          <tr>
            <th>Progetto</th>
            <th>Cliente / Filiale</th>
            <th>{"Provincia" if mode=="branch" else "Filiale"}</th>
            <th>MWp</th>
            <th>Fase</th>
            <th>Stato</th>
            <th>Avanz.</th>
            <th>Ultimo aggiornamento</th>
          </tr>
        </thead>
        <tbody id="projectBody"></tbody>
      </table>
    </div>
    {branch_back_footer}
  </div>

  <div class="gantt-shell">
    <div class="table-head">
      <div class="table-title">Gantt chart</div>
      <div class="small">La vista segue gli stessi filtri della sezione sopra.</div>
    </div>
    <div class="gantt-scroll">
      <div class="gantt" id="ganttWrap"></div>
    </div>
  </div>
</div>

<script>
const PAGE_MODE = {json.dumps(mode)};
const PAGE_BRANCH = {json.dumps(branch_name, ensure_ascii=False)};
const PAGE_AREA = {json.dumps(branch_area, ensure_ascii=False)};
const data = {page_data};
const branchDirectory = {branches_data};

const colors = {{
  'Studio di fattibilità':'var(--violet)',
  'Progettazione':'var(--brown)',
  'Autorizzazioni':'var(--gold)',
  'Procurement':'var(--blue)',
  'Costruzione':'var(--green)',
  'Commissioning':'var(--olive)',
  'Da definire':'var(--neutral)'
}};
const filters = {{
  area: document.getElementById('areaFilter'),
  scope: document.getElementById('{scope_id}'),
  fase: document.getElementById('faseFilter'),
  state: document.getElementById('stateFilter')
}};

function fmtNum(v) {{
  return new Intl.NumberFormat('it-IT', {{minimumFractionDigits:1, maximumFractionDigits:1}}).format(Number(v || 0));
}}
function phaseChip(fase) {{
  if (fase === 'Procurement') return 'proc';
  if (fase === 'Costruzione') return 'costr';
  if (fase === 'Commissioning') return 'comm';
  if (fase === 'Studio di fattibilità') return 'std';
  if (fase === 'Progettazione') return 'prog';
  return 'aut';
}}
function stateChip(stato) {{
  if (stato === 'Attivo') return 'state-attivo';
  if (stato === 'Completato') return 'state-completato';
  return 'state-pipeline';
}}
function fillSelect(el, values, placeholder, selected='') {{
  if (!el) return;
  el.innerHTML = `<option value="">${{placeholder}}</option>` + values.map(v => `<option ${{v===selected?'selected':''}}>${{v}}</option>`).join('');
}}
function uniqueSorted(arr) {{
  return [...new Set(arr.filter(Boolean))].sort((a,b)=>a.localeCompare(b, 'it'));
}}
const phaseValues = uniqueSorted(data.map(d=>d.fase));
const stateValues = uniqueSorted(data.map(d=>d.stato));
fillSelect(filters.fase, phaseValues, 'Tutte le fasi');
fillSelect(filters.state, stateValues, 'Tutti gli stati');

if (PAGE_MODE === 'global') {{
  fillSelect(filters.area, uniqueSorted(data.map(d=>d.area)), 'Tutte le aree');
  syncScopeByArea();
}} else {{
  fillSelect(filters.scope, uniqueSorted(data.map(d=>d.provincia)), 'Tutte le province');
}}

function syncScopeByArea() {{
  if (PAGE_MODE !== 'global') return;
  const prev = filters.scope.value;
  const allowed = uniqueSorted(data.filter(d => !filters.area.value || d.area === filters.area.value).map(d => d.filiale));
  fillSelect(filters.scope, allowed, 'Tutte le filiali', allowed.includes(prev) ? prev : '');
}}
if (filters.area) filters.area.addEventListener('change', () => {{ syncScopeByArea(); render(); }});

function getFiltered() {{
  return data.filter(d =>
    (!filters.fase.value || d.fase === filters.fase.value) &&
    (!filters.state.value || d.stato === filters.state.value) &&
    (PAGE_MODE !== 'global' || !filters.area.value || d.area === filters.area.value) &&
    (PAGE_MODE === 'global'
      ? (!filters.scope.value || d.filiale === filters.scope.value)
      : (!filters.scope.value || d.provincia === filters.scope.value))
  );
}}
function updateJumpMenu() {{
  if (PAGE_MODE !== 'global') return;
  const select = document.getElementById('jumpBranch');
  if (!select) return;
  select.innerHTML = `<option value="">Seleziona una filiale</option>` +
    branchDirectory.map(b => `<option value="${{b.slug}}">${{b.name}}</option>`).join('');
  document.getElementById('jumpBtn').addEventListener('click', () => {{
    if (select.value) window.location.href = `filiali/${{select.value}}.html`;
  }});
}}
function renderBars(rows) {{
  const groupKey = PAGE_MODE === 'global' ? 'filiale' : 'provincia';
  const grouped = Object.entries(rows.reduce((acc, r) => {{
    const key = r[groupKey] || '';
    if (!key) return acc;
    acc[key] = (acc[key] || 0) + Number(r.mwp || 0);
    return acc;
  }}, {{}})).sort((a,b)=>b[1]-a[1]);
  const max = Math.max(1, ...grouped.map(x=>x[1]), 1);
  const bars = document.getElementById('scopeBars');
  bars.innerHTML = grouped.map(([label, val]) => {{
    const branch = PAGE_MODE === 'global' ? branchDirectory.find(b => b.name === label) : null;
    const labelHtml = branch
      ? `<a class="bar-label-link" href="filiali/${{branch.slug}}.html">${{label}}</a>`
      : `<span>${{label}}</span>`;
    const wrapStart = branch ? `<div class="bar-row linkable" onclick="window.location.href='filiali/${{branch.slug}}.html'">` : `<div class="bar-row">`;
    return `${{wrapStart}}${{labelHtml}}<div class="bar-track"><div class="bar-fill" style="width:${{(val/max)*100}}%"></div></div><div><strong>${{fmtNum(val)}}</strong></div></div>`;
  }}).join('');
  return grouped;
}}
function renderDonut(rows) {{
  const phaseCounts = uniqueSorted(rows.map(d=>d.fase)).map(p => [p, rows.filter(d=>d.fase===p).length]).filter(x=>x[1]>0);
  const sum = Math.max(1, phaseCounts.reduce((a,b)=>a+b[1],0));
  let curr = 0;
  const stops = phaseCounts.map(([p,c]) => {{
    const start = curr / sum * 100;
    curr += c;
    const end = curr / sum * 100;
    return `${{colors[p] || 'var(--gold)'}} ${{start}}% ${{end}}%`;
  }});
  document.getElementById('donut').style.background = `conic-gradient(${{stops.join(',')}})`;
  document.getElementById('phaseLegend').innerHTML = phaseCounts.map(([p,c]) =>
    `<div class="legend-row"><span class="dot" style="background:${{colors[p] || 'var(--gold)'}}"></span><span>${{p}}</span><strong>${{c}}</strong></div>`
  ).join('');
  return phaseCounts;
}}
function renderTable(rows) {{
  const third = PAGE_MODE === 'global' ? 'filiale' : 'provincia';
  document.getElementById('projectBody').innerHTML = rows.map(r => `
    <tr>
      <td><strong>${{r.progetto || ''}}</strong></td>
      <td>${{r.cliente || r.filiale || ''}}</td>
      <td>${{r[third] || ''}}</td>
      <td><strong>${{fmtNum(r.mwp)}}</strong></td>
      <td><span class="chip ${{phaseChip(r.fase)}}">${{r.fase || ''}}</span></td>
      <td><span class="chip ${{stateChip(r.stato)}}">${{r.stato || ''}}</span></td>
      <td><div class="progress"><span>${{Math.round((r.avanz || 0) * 100)}}%</span><div class="track"><div class="fill" style="width:${{Math.round((r.avanz || 0) * 100)}}%"></div></div></div></td>
      <td>${{r.ultimo_agg || ''}}</td>
    </tr>
  `).join('');
}}
function monthSeries(rows) {{
  const dates = [];
  rows.forEach(r => (r.segments || []).forEach(s => {{
    if (s.start) dates.push(new Date(s.start));
    if (s.end) dates.push(new Date(s.end));
  }}));
  if (!dates.length) {{
    const now = new Date();
    return [new Date(now.getFullYear(), now.getMonth(), 1)];
  }}
  let min = new Date(Math.min(...dates.map(d => d.getTime())));
  let max = new Date(Math.max(...dates.map(d => d.getTime())));
  min = new Date(min.getFullYear(), min.getMonth(), 1);
  max = new Date(max.getFullYear(), max.getMonth(), 1);
  const months = [];
  const cursor = new Date(min);
  while (cursor <= max) {{
    months.push(new Date(cursor));
    cursor.setMonth(cursor.getMonth() + 1);
  }}
  return months;
}}
function monthName(d) {{
  return d.toLocaleDateString('it-IT', {{month:'short', year:'numeric'}});
}}
function renderGantt(rows) {{
  const el = document.getElementById('ganttWrap');
  const rowsWithSeg = rows.filter(r => (r.segments || []).length);
  if (!rowsWithSeg.length) {{
    el.innerHTML = '<div class="gantt-empty">Nessuna data disponibile per costruire la Gantt con i filtri correnti.</div>';
    return;
  }}
  const months = monthSeries(rowsWithSeg);
  const monthStart = months[0];
  const monthEnd = new Date(months[months.length - 1].getFullYear(), months[months.length - 1].getMonth() + 1, 1);
  const total = Math.max(1, monthEnd - monthStart);

  let html = `
    <div class="gantt-header">
      <div class="gantt-label">Progetto</div>
      <div class="gantt-timeline" style="--months:${{months.length}}">
        ${{months.map(m => `<div class="gantt-month">${{monthName(m)}}</div>`).join('')}}
      </div>
    </div>
  `;
  html += rowsWithSeg.map(r => {{
    const segs = (r.segments || []).map(s => {{
      const start = new Date(s.start);
      const end = new Date(s.end);
      const left = ((start - monthStart) / total) * 100;
      const width = Math.max(1.8, ((new Date(end.getFullYear(), end.getMonth()+1, 1) - start) / total) * 100);
      return `<div class="gantt-seg" title="${{s.label}}: ${{s.start}} → ${{s.end}}" style="left:${{left}}%;width:${{width}}%;background:${{colors[s.label] || 'var(--gold)'}}"></div>`;
    }}).join('');
    return `
      <div class="gantt-row">
        <div>
          <div class="gantt-name">${{r.progetto || ''}}</div>
          <div class="gantt-sub">${{r.cliente || r.filiale || ''}}</div>
        </div>
        <div class="gantt-track" style="--months:${{months.length}}">${{segs}}</div>
      </div>
    `;
  }}).join('');
  el.innerHTML = html;
}}
function render() {{
  const rows = getFiltered();
  const totalProjects = rows.length;
  const totalMwp = rows.reduce((a,b)=>a + Number(b.mwp || 0), 0);
  const buildCount = rows.filter(d=>d.fase === 'Costruzione').length;
  const agriCount = rows.filter(d => (d.note || '').toLowerCase().includes('agrivoltaico')).length;

  document.getElementById('kpiProjects').textContent = totalProjects;
  document.getElementById('kpiMwp').textContent = fmtNum(totalMwp);
  document.getElementById('kpiBuild').textContent = buildCount;
  document.getElementById('kpiAgri').textContent = agriCount;
  document.getElementById('donutTotal').textContent = totalProjects;

  const phaseCounts = renderDonut(rows);
  const grouped = renderBars(rows);
  renderTable(rows);
  renderGantt(rows);

  const topGroup = grouped[0];
  const topPhase = phaseCounts.slice().sort((a,b)=>b[1]-a[1])[0];
  const groupLabel = PAGE_MODE === 'global' ? 'La filiale più pesante è' : 'La provincia più pesante è';
  document.getElementById('insightBox').innerHTML =
    `<strong>Lettura rapida.</strong> ${{totalProjects}} progetti per <strong>${{fmtNum(totalMwp)}} MWp</strong>. ` +
    (topGroup ? `${{groupLabel}} <strong>${{topGroup[0]}}</strong> con <strong>${{fmtNum(topGroup[1])}} MWp</strong>. ` : '') +
    (topPhase ? `La fase dominante resta <strong>${{topPhase[0]}}</strong> con <strong>${{topPhase[1]}}</strong> progetti.` : '');
}}

[filters.scope, filters.fase, filters.state].forEach(el => el && el.addEventListener('change', render));
updateJumpMenu();
render();
</script>
</body>
</html>"""

def main():
    cfg = load_config()
    agg_dir = Path(cfg["excel_aggregatore_dir"])
    filiali_dir = Path(cfg["excel_filiali_dir"])
    repo_root = Path(cfg["repo_root_dir"])
    docs_dir = Path(cfg["docs_dir"])
    docs_filiali = docs_dir / "filiali"
    docs_filiali.mkdir(parents=True, exist_ok=True)
    agg_file = agg_dir / cfg["aggregatore_filename"]

    global_rows, _, _ = read_sheet_rows(agg_file)

    branch_rows = {}
    branch_meta = {}
    for fp in sorted(filiali_dir.glob("Fotovoltaico_*.xlsx")):
        if fp.name == cfg["aggregatore_filename"]:
            continue
        rows, bname, barea = read_sheet_rows(fp)
        name = bname or fp.stem.replace("Fotovoltaico_", "")
        branch_rows[name] = rows
        branch_meta[name] = {"area": barea, "slug": slugify(name)}

    for r in global_rows:
        b = r["filiale"]
        if b and b not in branch_rows:
            branch_rows[b] = [x for x in global_rows if x["filiale"] == b]
            branch_meta[b] = {"area": r["area"], "slug": slugify(b)}

    all_branches = [{"name": name, "slug": meta["slug"], "area": meta["area"]} for name, meta in sorted(branch_meta.items())]

    docs_dir.mkdir(parents=True, exist_ok=True)
    (repo_root / ".nojekyll").write_text("", encoding="utf-8")

    (docs_dir / "index.html").write_text(
        render_page(
            "Pipeline Fotovoltaico - Overview nazionale",
            "Dashboard globale.",
            "global",
            global_rows,
            all_branches=all_branches
        ),
        encoding="utf-8"
    )

    for old in docs_filiali.glob("*.html"):
        old.unlink()

    for name, rows in sorted(branch_rows.items()):
        area = branch_meta.get(name, {}).get("area", "")
        slug = branch_meta.get(name, {}).get("slug", slugify(name))
        (docs_filiali / f"{slug}.html").write_text(
            render_page(
                f"Pipeline Fotovoltaico - Filiale {name}",
                "Dashboard filiale.",
                "branch",
                rows,
                branch_name=name,
                branch_area=area,
                all_branches=all_branches
            ),
            encoding="utf-8"
        )

    print("Sito aggiornato correttamente.")
    print(f"Overview: {docs_dir / 'index.html'}")
    print(f"Filiali generate: {len(branch_rows)}")

if __name__ == "__main__":
    main()
